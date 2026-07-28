"""Stage 1: build the canonical benefit list from the Topic Tree Excel file.

Reads the 'Raw Data' tab, dedupes column B (FULL_NAME_TEXT), filters out
procedure codes, and attaches a normalized matching key plus associated
TOPIC_IDs / tree paths for each retained benefit name.

Usage:
    python3 build_benefits.py <topic_tree.xlsx> <output_dir>
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from normalize import normalized_key

# A value is a procedure code if it starts with a 5-character alphanumeric
# code followed by a dash (spacing around the dash is flexible to catch
# entries missing a space, e.g. "G0296 -Counseling visit...").
CODE_PATTERN = re.compile(r"^[A-Za-z0-9]{5}\s*-\s*")

# Many retained (non-procedure-code) benefit names are still prefixed with a
# 3-5 digit internal reference/revenue code (e.g. "0547 - Ambulance pharmacy",
# "2103 - Massage") -- confirmed via real corpus testing that these codes
# never appear verbatim next to the benefit description in the actual PDF
# text, so leaving them in normalized_key makes ~28% of the Topic Tree
# structurally unmatchable (556/1986 benefits, only 2 with any mention at
# all, and both of those were coincidental fuzzy-matching noise against
# unrelated nearby numbers, not the code itself). Stripped only for the
# matching key -- benefit_name keeps the code for display/traceability.
REVENUE_CODE_PREFIX = re.compile(r"^\d{3,5}\s*-\s*")


def is_procedure_code(name):
    return bool(CODE_PATTERN.match(name))


def strip_code_prefix(name):
    return REVENUE_CODE_PREFIX.sub("", name, count=1)


def build_benefits(xlsx_path, sheet_name="Raw Data"):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = ws.iter_rows(min_row=2, values_only=True)
    header = None
    ws2 = wb[sheet_name]
    header_row = next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {name: i for i, name in enumerate(header_row) if name is not None}

    required = ["TOPIC_ID", "FULL_NAME_TEXT", "PRIMARY_PATH_TEXT"]
    missing = [c for c in required if c not in col_index]
    if missing:
        raise ValueError(f"Missing expected columns in '{sheet_name}': {missing}")

    by_name = {}
    total_rows = 0
    skipped_codes = 0
    skipped_blank = 0

    for row in rows:
        total_rows += 1
        name = row[col_index["FULL_NAME_TEXT"]]
        if name is None or not str(name).strip():
            skipped_blank += 1
            continue
        name = str(name).strip()

        if is_procedure_code(name):
            skipped_codes += 1
            continue

        topic_id = row[col_index["TOPIC_ID"]]
        path = row[col_index["PRIMARY_PATH_TEXT"]]

        entry = by_name.setdefault(
            name,
            {
                "benefit_name": name,
                "normalized_key": normalized_key(strip_code_prefix(name)),
                "topic_ids": [],
                "tree_paths": [],
            },
        )
        if topic_id is not None and topic_id not in entry["topic_ids"]:
            entry["topic_ids"].append(topic_id)
        if path is not None and path not in entry["tree_paths"]:
            entry["tree_paths"].append(path)

    benefits = sorted(by_name.values(), key=lambda e: e["benefit_name"].lower())

    stats = {
        "total_rows": total_rows,
        "skipped_blank": skipped_blank,
        "skipped_procedure_codes": skipped_codes,
        "unique_benefits": len(benefits),
    }
    return benefits, stats


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--sheet", default="Raw Data")
    args = parser.parse_args()

    benefits, stats = build_benefits(args.xlsx_path, args.sheet)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    out_path = args.output_dir / "benefits.json"
    with open(out_path, "w") as f:
        json.dump({"stats": stats, "benefits": benefits}, f, indent=2, default=str)

    print(f"Wrote {len(benefits)} benefits to {out_path}", file=sys.stderr)
    print(json.dumps(stats, indent=2), file=sys.stderr)


if __name__ == "__main__":
    main()
